"""
FlowERP — Dynamic Reports Engine
==================================
Endpoints:
  GET  /reports                     list saved reports
  POST /reports                     create saved report
  POST /reports/run                 run a report ad-hoc
  GET  /reports/{id}                get saved report
  PUT  /reports/{id}                update saved report
  DELETE /reports/{id}              delete saved report
  GET  /reports/{id}/run            run a saved report
  GET  /reports/{id}/export/excel   export to Excel
  GET  /reports/{id}/export/pdf     export to PDF
  GET  /reports/modules             list available modules and fields
"""
import io
import math
from datetime import datetime, timezone
from typing import Optional, List, Any
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from pydantic import BaseModel
import sqlalchemy as sa
from sqlalchemy.dialects.postgresql import UUID as PG_UUID, JSON

from app.core.database import get_db, Base
from app.core.dependencies import get_current_active_user
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["Reports"])


# ── Saved Report model ─────────────────────────────────────────────────
class SavedReport(Base):
    __tablename__ = "saved_reports"
    tenant_id   = sa.Column(PG_UUID(as_uuid=True), sa.ForeignKey("tenants.id", ondelete="CASCADE"), nullable=False, index=True)
    name        = sa.Column(sa.String(300), nullable=False)
    description = sa.Column(sa.Text, nullable=True)
    module      = sa.Column(sa.String(100), nullable=False)
    config      = sa.Column(JSON, nullable=False, default=dict)
    is_public   = sa.Column(sa.Boolean, default=False, nullable=False)
    schedule    = sa.Column(sa.String(50), nullable=True)
    last_run_at = sa.Column(sa.DateTime(timezone=True), nullable=True)
    created_by  = sa.Column(PG_UUID(as_uuid=True), sa.ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    deleted_at  = sa.Column(sa.DateTime(timezone=True), nullable=True)


# ── Schemas ────────────────────────────────────────────────────────────
class ReportFilter(BaseModel):
    field:    str
    operator: str    # equals, not_equals, gt, lt, gte, lte, contains, in, between, is_null
    value:    Any = None
    value2:   Any = None   # for between


class ReportConfig(BaseModel):
    module:       str
    fields:       List[str] = []
    filters:      List[ReportFilter] = []
    sort_by:      Optional[str] = None
    sort_dir:     str = "desc"
    limit:        int = 1000
    group_by:     Optional[str] = None


class ReportCreate(BaseModel):
    name:        str
    description: Optional[str] = None
    module:      str
    config:      dict = {}
    is_public:   bool = False
    schedule:    Optional[str] = None


class ReportUpdate(BaseModel):
    name:        Optional[str] = None
    description: Optional[str] = None
    config:      Optional[dict] = None
    is_public:   Optional[bool] = None
    schedule:    Optional[str] = None


# ── Available modules & fields ─────────────────────────────────────────
REPORT_MODULES = {
    "sales_orders": {
        "label":  "Sales Orders",
        "table":  "sales_orders",
        "fields": [
            {"key": "order_number",   "label": "Order #",        "type": "string"},
            {"key": "customer_name",  "label": "Customer",        "type": "string"},
            {"key": "total_amount",   "label": "Total Amount",    "type": "number"},
            {"key": "subtotal",       "label": "Subtotal",        "type": "number"},
            {"key": "tax_amount",     "label": "Tax Amount",      "type": "number"},
            {"key": "paid_amount",    "label": "Paid Amount",     "type": "number"},
            {"key": "balance_due",    "label": "Balance Due",     "type": "number"},
            {"key": "status",         "label": "Status",          "type": "string"},
            {"key": "payment_status", "label": "Payment Status",  "type": "string"},
            {"key": "order_date",     "label": "Order Date",      "type": "date"},
            {"key": "delivery_date",  "label": "Delivery Date",   "type": "date"},
            {"key": "created_at",     "label": "Created",         "type": "datetime"},
        ],
    },
    "crm_leads": {
        "label":  "CRM Leads",
        "table":  "crm_leads",
        "fields": [
            {"key": "name",            "label": "Company",        "type": "string"},
            {"key": "contact_name",    "label": "Contact",        "type": "string"},
            {"key": "email",           "label": "Email",          "type": "string"},
            {"key": "source",          "label": "Source",         "type": "string"},
            {"key": "status",          "label": "Status",         "type": "string"},
            {"key": "estimated_value", "label": "Est. Value",     "type": "number"},
            {"key": "created_at",      "label": "Created",        "type": "datetime"},
        ],
    },
    "inventory_products": {
        "label":  "Inventory",
        "table":  "inventory_products",
        "fields": [
            {"key": "name",          "label": "Product Name",   "type": "string"},
            {"key": "sku",           "label": "SKU",            "type": "string"},
            {"key": "category",      "label": "Category",       "type": "string"},
            {"key": "stock",         "label": "Stock",          "type": "number"},
            {"key": "reorder_point", "label": "Reorder Point",  "type": "number"},
            {"key": "cost_price",    "label": "Cost Price",     "type": "number"},
            {"key": "selling_price", "label": "Selling Price",  "type": "number"},
            {"key": "status",        "label": "Status",         "type": "string"},
        ],
    },
    "purchase_orders": {
        "label":  "Purchase Orders",
        "table":  "purchase_orders",
        "fields": [
            {"key": "order_number",   "label": "PO #",            "type": "string"},
            {"key": "vendor_name",    "label": "Vendor",          "type": "string"},
            {"key": "total_amount",   "label": "Total Amount",    "type": "number"},
            {"key": "status",         "label": "Status",          "type": "string"},
            {"key": "payment_status", "label": "Payment Status",  "type": "string"},
            {"key": "order_date",     "label": "Order Date",      "type": "date"},
            {"key": "expected_date",  "label": "Expected Date",   "type": "date"},
        ],
    },
    "hr_employees": {
        "label":  "Employees",
        "table":  "hr_employees",
        "fields": [
            {"key": "name",        "label": "Name",         "type": "string"},
            {"key": "email",       "label": "Email",        "type": "string"},
            {"key": "department",  "label": "Department",   "type": "string"},
            {"key": "designation", "label": "Designation",  "type": "string"},
            {"key": "role",        "label": "Role",         "type": "string"},
            {"key": "status",      "label": "Status",       "type": "string"},
            {"key": "join_date",   "label": "Join Date",    "type": "date"},
        ],
    },
    "invoices": {
        "label":  "Invoices",
        "table":  "invoices",
        "fields": [
            {"key": "customer_name", "label": "Customer",       "type": "string"},
            {"key": "amount",        "label": "Amount",         "type": "number"},
            {"key": "paid",          "label": "Paid",           "type": "number"},
            {"key": "due",           "label": "Balance Due",    "type": "number"},
            {"key": "status",        "label": "Status",         "type": "string"},
            {"key": "issue_date",    "label": "Issue Date",     "type": "date"},
            {"key": "due_date",      "label": "Due Date",       "type": "date"},
        ],
    },
}

OPERATORS = {
    "equals":      lambda col, v, v2: col == v,
    "not_equals":  lambda col, v, v2: col != v,
    "gt":          lambda col, v, v2: col > v,
    "lt":          lambda col, v, v2: col < v,
    "gte":         lambda col, v, v2: col >= v,
    "lte":         lambda col, v, v2: col <= v,
    "contains":    lambda col, v, v2: col.ilike(f"%{v}%"),
    "starts_with": lambda col, v, v2: col.ilike(f"{v}%"),
    "is_null":     lambda col, v, v2: col.is_(None),
    "not_null":    lambda col, v, v2: col.isnot(None),
    "in":          lambda col, v, v2: col.in_(v if isinstance(v, list) else [v]),
}


def report_out(r: SavedReport) -> dict:
    return {
        "id": str(r.id), "name": r.name, "description": r.description,
        "module": r.module, "config": r.config, "is_public": r.is_public,
        "schedule": r.schedule,
        "last_run_at": r.last_run_at.isoformat() if r.last_run_at else None,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    }


# ── Module/field discovery ─────────────────────────────────────────────
@router.get("/modules")
async def list_modules(current_user: User = Depends(get_current_active_user)):
    return {"modules": [{"id": k, **{kk: vv for kk, vv in v.items() if kk != "table"}} for k, v in REPORT_MODULES.items()]}


# ── Saved report CRUD ──────────────────────────────────────────────────
@router.get("")
async def list_reports(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(SavedReport).where(
        SavedReport.tenant_id == current_user.tenant_id,
        SavedReport.deleted_at.is_(None),
    ).order_by(SavedReport.created_at.desc())
    items = (await db.execute(q)).scalars().all()
    return {"items": [report_out(r) for r in items], "total": len(items)}


@router.post("", status_code=201)
async def create_report(
    payload: ReportCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    r = SavedReport(tenant_id=current_user.tenant_id, created_by=current_user.id, **payload.model_dump())
    db.add(r); await db.flush()
    return report_out(r)


@router.get("/{report_id}")
async def get_report(
    report_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(SavedReport).where(SavedReport.id == report_id, SavedReport.tenant_id == current_user.tenant_id))
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Report not found")
    return report_out(r)


@router.put("/{report_id}")
async def update_report(
    report_id: UUID,
    payload: ReportUpdate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(SavedReport).where(SavedReport.id == report_id, SavedReport.tenant_id == current_user.tenant_id))
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Report not found")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    return report_out(r)


@router.delete("/{report_id}")
async def delete_report(
    report_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(SavedReport).where(SavedReport.id == report_id, SavedReport.tenant_id == current_user.tenant_id))
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Report not found")
    r.deleted_at = datetime.now(timezone.utc)
    return {"success": True}


# ── Run engine ─────────────────────────────────────────────────────────
async def _run_report_query(config: dict, tenant_id, db: AsyncSession) -> list[dict]:
    """Execute a report config against the database and return rows."""
    module_key = config.get("module", "sales_orders")
    module     = REPORT_MODULES.get(module_key)
    if not module:
        raise HTTPException(400, f"Unknown module: {module_key}")

    table_name = module["table"]
    fields     = config.get("fields", [f["key"] for f in module["fields"]])
    filters    = config.get("filters", [])
    sort_by    = config.get("sort_by") or "created_at"
    sort_dir   = config.get("sort_dir", "desc")
    limit      = min(int(config.get("limit", 1000)), 5000)

    # Build SELECT columns
    valid_field_keys = {f["key"] for f in module["fields"]}
    select_cols = [f for f in fields if f in valid_field_keys]
    if not select_cols:
        select_cols = [f["key"] for f in module["fields"]]

    cols_sql = ", ".join(f'"{c}"' for c in select_cols)

    # Build WHERE clauses
    where_parts = [f'"tenant_id" = \'{tenant_id}\'']
    where_parts.append('"deleted_at" IS NULL')

    for flt in filters:
        field = flt.get("field", "")
        op    = flt.get("operator", "equals")
        value = flt.get("value")

        if field not in valid_field_keys:
            continue
        if value is None and op not in ("is_null", "not_null"):
            continue

        if op == "equals":
            where_parts.append(f'"{field}" = \'{value}\'')
        elif op == "not_equals":
            where_parts.append(f'"{field}" != \'{value}\'')
        elif op == "gt":
            where_parts.append(f'CAST("{field}" AS NUMERIC) > {float(value)}')
        elif op == "lt":
            where_parts.append(f'CAST("{field}" AS NUMERIC) < {float(value)}')
        elif op == "gte":
            where_parts.append(f'CAST("{field}" AS NUMERIC) >= {float(value)}')
        elif op == "lte":
            where_parts.append(f'CAST("{field}" AS NUMERIC) <= {float(value)}')
        elif op == "contains":
            where_parts.append(f'"{field}" ILIKE \'%{value}%\'')
        elif op == "is_null":
            where_parts.append(f'"{field}" IS NULL')
        elif op == "not_null":
            where_parts.append(f'"{field}" IS NOT NULL')

    where_sql = " AND ".join(where_parts)
    sort_dir_sql = "DESC" if sort_dir.lower() == "desc" else "ASC"
    sort_col = sort_by if sort_by in valid_field_keys else "created_at"

    sql = f"""
        SELECT {cols_sql}
        FROM {table_name}
        WHERE {where_sql}
        ORDER BY "{sort_col}" {sort_dir_sql}
        LIMIT {limit}
    """

    try:
        result = await db.execute(text(sql))
        rows = result.fetchall()
        return [dict(zip(select_cols, row)) for row in rows]
    except Exception as e:
        raise HTTPException(500, f"Query failed: {str(e)}")


@router.post("/run")
async def run_report_adhoc(
    config: ReportConfig,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Run a report ad-hoc without saving it."""
    rows = await _run_report_query(config.model_dump(), current_user.tenant_id, db)
    return {
        "rows": rows,
        "total": len(rows),
        "module": config.module,
        "fields": config.fields,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@router.get("/{report_id}/run")
async def run_saved_report(
    report_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Run a saved report and update last_run_at."""
    result = await db.execute(
        select(SavedReport).where(SavedReport.id == report_id, SavedReport.tenant_id == current_user.tenant_id)
    )
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Report not found")

    rows = await _run_report_query(r.config, current_user.tenant_id, db)
    r.last_run_at = datetime.now(timezone.utc)

    return {
        "report_id": str(r.id),
        "report_name": r.name,
        "rows": rows,
        "total": len(rows),
        "generated_at": r.last_run_at.isoformat(),
    }


# ── Excel export ───────────────────────────────────────────────────────
@router.get("/{report_id}/export/excel")
async def export_report_excel(
    report_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SavedReport).where(SavedReport.id == report_id, SavedReport.tenant_id == current_user.tenant_id)
    )
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Report not found")

    rows = await _run_report_query(r.config, current_user.tenant_id, db)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = r.name[:31]

        # Header row
        if rows:
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F46E5")
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1):
                    ws.cell(row=ri, column=ci, value=v)

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{r.name.replace(' ', '_')}.xlsx"
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        # Fall back to CSV
        import csv
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{r.name}.csv"'},
        )


# ── Dashboard summary endpoint ─────────────────────────────────────────
@router.get("/dashboard/summary")
async def dashboard_summary(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_id = str(current_user.tenant_id)
    try:
        # Run lightweight queries
        orders_q = await db.execute(text(f"SELECT COUNT(*), COALESCE(SUM(total_amount),0) FROM sales_orders WHERE tenant_id='{tenant_id}' AND deleted_at IS NULL"))
        orders_row = orders_q.fetchone()

        products_q = await db.execute(text(f"SELECT COUNT(*) FROM inventory_products WHERE tenant_id='{tenant_id}' AND deleted_at IS NULL"))
        products_count = products_q.scalar()

        leads_q = await db.execute(text(f"SELECT COUNT(*) FROM crm_leads WHERE tenant_id='{tenant_id}' AND deleted_at IS NULL"))
        leads_count = leads_q.scalar()

        return {
            "total_orders":    int(orders_row[0] or 0),
            "monthly_revenue": float(orders_row[1] or 0),
            "active_products": int(products_count or 0),
            "total_leads":     int(leads_count or 0),
        }
    except Exception:
        return {"total_orders": 0, "monthly_revenue": 0, "active_products": 0, "total_leads": 0}
